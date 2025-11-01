import time
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
from datetime import datetime
import threading
import pyaudio
import wave  # For saving WAV file
import whisper  # OpenAI Whisper for transcription
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pyttsx3
from tkinter import messagebox
import re
import tempfile
import customtkinter as ctk  # New import for modern UI
import platform
import subprocess
from PIL import Image


def safe_json_loads(data):
    try:
        return json.loads(data)
    except Exception:
        return {}


# Set CustomTkinter appearance (modern look)
ctk.set_appearance_mode("light")  # Modes: "light", "dark", "system"
ctk.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue"


class WeldingShopApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Al Tasnim Enterprises LLC")
        self.root.geometry("1400x900")  # Wider and taller for better table visibility
        
        self.entry_frames = {}      # map field_id -> container frame
        self.mic_buttons = {}       # map field_id -> mic button widget
        self.display_labels = {}
        # Data storage
        self.records = []  # List of dicts for table rows
        self.data_file = "welding_data.json"
        self.load_data()
        
        self.header_entries = {}
        self.table_entries = {}
        self.signature_entries = {}
        self.status_label = None

        # Header fields data
        self.header_data = {
            "contract_number": "",
            "contract_title": "",
            "report_number": "",
            "date": datetime.now().strftime("%Y-%m-%d"),  # Activity Date
            "po_wo_number": "",
            "client_wps_number": "",
            "project_title_wellID": "",
            "drawing_no": "",
            "line_no": "",
            "site_name": "",
            "job_desc": "",
            "location": "",
            # Welding Consumable
            "aws_classification": "",
            "electrode_dia": "",
            "manufacturer_batch": "",
            # Signatures
            "permit_holder_name": "",
            "permit_holder_signature": "",
            "permit_holder_date": "",
            "qci_name": "",
            "qci_signature": "",
            "qci_date": "",
            "pdo_name": "",
            "pdo_signature": "",
            "pdo_date": "",
            "data_entry_name": "",
            "data_entry_signature": "",
            "data_entry_date": "",
        }

        # Language settings
        self.current_lang = "en"
        self.translations = {
            "en": {
                "title": "AL TASNIM ENTERPRISES LLC",
                "form_title": "Daily Welding Production - Visual Inspection Report",
                "contract_number": "Contract No.:",
                "contract_title": "Contract Title:",
                "report_number": "Report No.:",
                "date": "Activity Date:",
                "po_wo_number": "PO / WO No.:",
                "client_wps_number": "Client WPS No.:",
                "project_title_wellID": "Project Title/Well ID:",
                "drawing_no": "Drawing/ISO No.:",
                "line_no": "Line No.:",
                "site_name": "Site Name:",
                "job_desc": "Job Description:",
                "location": "Location:",
                "table_headers": {
                    "sr_no": "Sr.No",
                    "kp_sec": "KP Sec.",
                    "weld_id": "Weld ID",
                    "wps_no": "WPS No.",
                    "material_gr_heat": "Material Gr./ Heat No.",
                    "size": "Size",
                    "thk": "Thk.",
                    "weld_side": "Weld Side",
                    "welder_process": "Welder No. / Welding Process",
                    "visual_i": "I",
                    "visual_ii": "II",
                    "root_hot": "Root/Hot",
                    "fill1": "Fill…………..",
                    "fill2": "Fill…………..",
                    "cap": "Cap…………..",
                    "final": "Final",
                    "fit_up": "Fit up",
                    "mtrl_comb": "Mtrl. Comb.",
                    "pipe_line": "Pipe Line",
                    "pipe_no": "Pipe No. / Spl. No.",
                    "pipe_length": "Pipe length (mtrs)",
                    "remarks": "Remarks",
                },
                "welding_consumable": "Welding Consumable",
                "aws_classification": "AWS Classification",
                "electrode_dia": "Electrode Dia. (mm)",
                "manufacturer_batch": "Manufacturer & Batch No.",
                "material_grade_legend": "Material Grade Legend: 1 = A 106 Gr. B;  2 = A 105N;  3 = A 234 Gr WPB;  4 = ISO 3183- L245 (Gr. B);  5 = ISO 3183 L290 (X42);  6 = ISO 3183- L360 (X52); 7 = ISO 3183- L415 (X60);  8 = ISO 3183- L450 (X65);  9 = ISO 3183- L485 (X70);  10 = A312 TP-316L/316;  11 = A403 WP-316L/316;  12 = A182 F316L/316;  13 = B444 (UNS N06625);  14 = B424 (UNS N08825);  15 = B668 (UNS N8028);  16 = A790 (UNS S31803);  17 = A182 F51 (UNS S31803);  18 = A928Cl:1 (UNS 31803);  19 = A 815 (UNS S31803);  20 = A 333 Gr. 6;  21 = A 420 WPL 6;  22 = A350 LF 2;  23 = AISI 4130;  24 = A694 F42;  25 = A694 F 52;  26 = A694 F60;  27 = A694 F 65;  28 = A694 F 70;  29 = A240 Gr. 316;  30 = B564 (UNS N08825);  31 = A358 TP-316L/316;  32 = B366 Gr: WPNICMC; 33 = ISO 15590-2/3- L245 (Gr. B);  34  = ISO 15590-2/3 L290 (X42); 35 = ISO 15590-2/3 L360 (X52); 36 = ISO 15590-2/3 L415 (X60); 37 = ISO 15590-2/3 L450 (X65);  38 = ISO 15590-2/3 L485 (X70) ;    39 =  S275 JR    ; 40 =  S355 JR  ; 41 =  S355 NL / K2+N    42 = ASTM A790/A928 (UNS S32750/ S32760 )  ; 43 = ASTM A182 F53 (UNS S32760)  ;    44 = ASTM A815 (UNS S32760)      ",
                "welding_process_legend": "Welding Process Legend:  P1 = GTAW;  P2: = SMAW;  P3 = GMAW;  P4 = FCAW;  P5 = SAW           Weld Side based on the direction of welding:  A = 12 - 6 O'Clock Position;  B = 6 - 12 O'Clock Position",
                "permit_holder": "ATNM Permit Holder",
                "qci": "ATNM QCI",
                "pdo": "PDO",
                "data_entry": "Data Entry By",
                "name": "Name:",
                "signature": "Signature:",
                "date": "Date:",
                "add_entry": "Add Entry",  # Not used in new UI
                "clear_form": "Clear Form",
                "download_excel": "Submit",
                "language": "Language:",
                "theme": "Theme:",
                "records_title": "Records",  # Not used
                "delete": "Delete",  # Not used
                "success_add": "Entry added successfully!",
                "success_export": "Form exported successfully!",
                "error_fill": "Please fill required fields",
                "recording": "Recording... Speak now",
                "mic_tooltip": "Click to record voice",
                "confirm_text": "Is this correct: '{}'? (Yes to confirm and lock field, No to re-record)",
                "confirm_prompt": "You said {}. Do you confirm?",
                "no_hear": "I didn't hear you. Please say yes or no.",
                "no_catch": "I didn't catch that. Please say yes or no.",
                "okay_retry": "Okay, please say it again.",
            },
            "ar": {
                # Arabic translations (placeholder, expand as needed)
                "title": "إدارة ورشة اللحام",
                "form_title": "تقرير إنتاج اللحام اليومي - تقرير التفتيش البصري",
                "contract_number": "رقم العقد:",
                "contract_title": "عنوان العقد:",
                "report_number": "رقم التقرير:",
                "date": "تاريخ النشاط:",
                "po_wo_number": "رقم PO / WO:",
                "client_wps_number": "رقم WPS العميل:",
                "project_title_wellID": "عنوان المشروع / معرف البئر:",
                "drawing_no": "رقم الرسم/ISO:",
                "line_no": "رقم الخط:",
                "site_name": "اسم الموقع:",
                "job_desc": "وصف الوظيفة:",
                "location": "الموقع:",
                "table_headers": {
                    "sr_no": "الرقم التسلسلي",
                    "kp_sec": "قسم KP",
                    "weld_id": "رقم اللحام",
                    "wps_no": "رقم WPS",
                    "material_gr_heat": "درجة المادة / رقم الحرارة",
                    "size": "الحجم",
                    "thk": "السماكة",
                    "weld_side": "جانب اللحام",
                    "welder_process": "رقم اللحام / عملية اللحام",
                    "visual_i": "I",
                    "visual_ii": "II",
                    "root_hot": "الجذر/ساخن",
                    "fill1": "ملء…………..",
                    "fill2": "ملء…………..",
                    "cap": "غطاء…………..",
                    "final": "نهائي",
                    "fit_up": "تركيب",
                    "mtrl_comb": "تركيبة المادة",
                    "pipe_line": "خط الأنابيب",
                    "pipe_no": "رقم الأنبوب / رقم خاص",
                    "pipe_length": "طول الأنبوب (متر)",
                    "remarks": "ملاحظات",
                },
                "welding_consumable": "مستهلكات اللحام",
                "aws_classification": "تصنيف AWS",
                "electrode_dia": "قطر الإلكترود (مم)",
                "manufacturer_batch": "الشركة المصنعة & رقم الدفعة",
                "material_grade_legend": "أسطورة درجة المادة: ...",  # Translate full
                "welding_process_legend": "أسطورة عملية اللحام: ...",  # Translate full
                "permit_holder": "حامل تصريح ATNM",
                "qci": "QCI ATNM",
                "pdo": "PDO",
                "data_entry": "إدخال البيانات بواسطة",
                "name": "الاسم:",
                "signature": "التوقيع:",
                "date": "التاريخ:",
                "add_entry": "إضافة سجل",
                "clear_form": "مسح النموذج",
                "download_excel": "إرسال النموذج",
                "language": "اللغة:",
                "theme": "الثيم:",
                "records_title": "السجلات",
                "delete": "حذف",
                "success_add": "تمت إضافة السجل بنجاح!",
                "success_export": "تم تصدير النموذج بنجاح!",
                "error_fill": "يرجى ملء الحقول المطلوبة",
                "recording": "جاري التسجيل... تحدث الآن",
                "mic_tooltip": "انقر للتسجيل الصوتي",
                "confirm_text": "هل هذا صحيح: '{}'؟ (نعم للتأكيد وتثبيت الحقل، لا لإعادة التسجيل)",
                "confirm_prompt": "قلت {}. هل تؤكد؟",
                "no_hear": "لم أسمعك. يرجى القول نعم أو لا.",
                "no_catch": "لم أفهم ذلك. يرجى القول نعم أو لا.",
                "okay_retry": "حسنا، يرجى قوله مرة أخرى.",
            },
        }

        # Number word dictionaries for conversion
        self.number_dicts = {
            "en": {
                "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
                "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11, "twelve": 12,
                "thirteen": 13, "fourteen": 14, "fifteen": 15, "sixteen": 16,
                "seventeen": 17, "eighteen": 18, "nineteen": 19, "twenty": 20,
                "thirty": 30, "forty": 40, "fifty": 50, "sixty": 60, "seventy": 70,
                "eighty": 80, "ninety": 90,
            },
            "ar": {
                "صفر": 0, "واحد": 1, "اثنان": 2, "ثلاثة": 3, "أربعة": 4, "خمسة": 5,
                "ستة": 6, "سبعة": 7, "ثمانية": 8, "تسعة": 9, "عشرة": 10,
                "أحد عشر": 11, "اثنا عشر": 12, "ثلاثة عشر": 13, "أربعة عشر": 14,
                "خمسة عشر": 15, "ستة عشر": 16, "سبعة عشر": 17, "ثمانية عشر": 18,
                "تسعة عشر": 19, "عشرون": 20, "ثلاثون": 30, "أربعون": 40,
                "خمسون": 50, "ستون": 60, "سبعون": 70, "ثمانون": 80, "تسعون": 90,
            },
        }

        # Whisper model
        self.whisper_model = None
        self.load_whisper_model()

        # Voice recognition
        self.is_recording = False
        self.audio = None
        self.stream = None

        # UI elements dicts
        self.header_entries = {}
        self.table_entries = {}  # {row: {field: entry}}
        self.signature_entries = {}

        # Build UI
        self.create_ui()
        self.update_language()
        self.load_to_ui()  # Load saved data to UI

    def load_whisper_model(self):
        try:
            self.whisper_model = whisper.load_model("medium")
            print("Loaded Whisper model: medium")
        except Exception as e:
            print(f"Error loading Whisper model: {e}. Install with: pip install openai-whisper")
            self.whisper_model = None

    def words_to_digits(self, text, lang):
        if not text:
            return None
        cleaned = re.sub(r"\s+", "", text.strip())
        if re.match(r"^\d+$", cleaned):
            return cleaned
        if lang == "en":
            normalized = text.lower().strip()
        else:
            normalized = text.strip()
        words = re.split(r"\s+and\s+|and\s+|,\s*|\s+", normalized)
        words = [w.strip() for w in words if w.strip()]
        if not words:
            return None
        num_values = []
        single_digits_only = True
        for word in words:
            num = self.number_dicts[lang].get(word, None)
            if num is None:
                return None
            num_values.append(num)
            if num > 9:
                single_digits_only = False
        if not num_values:
            return None
        if single_digits_only:
            return "".join(str(num) for num in num_values)
        else:
            return str(sum(num_values))

    # -------------------- TTS helpers --------------------
    def speak_sync(self, text, timeout=20):
        """
        Speak text synchronously using pyttsx3 with language-appropriate voice selection.
        """
        try:
            local_tts = pyttsx3.init()
            local_tts.setProperty("rate", 160)
            local_tts.setProperty("volume", 1.0)  # Max volume

            # Select voice based on current language
            voices = local_tts.getProperty('voices')
            print("[DEBUG] Available TTS voices:")
            for v in voices:
                print(f" - Name: {v.name}, ID: {v.id}, Languages: {v.languages}")

            target_lang = 'arabic' if self.current_lang == "ar" else 'english'
            selected = False
            for voice in voices:
                langs = voice.languages if isinstance(voice.languages, list) else [str(voice.languages)]
                name_lower = voice.name.lower()
                lang_lower = [l.lower() for l in langs] if isinstance(langs, list) else [langs.lower()]
                if self.current_lang == "ar" and any('ar' in l or 'arabic' in l for l in lang_lower):
                    local_tts.setProperty('voice', voice.id)
                    selected = True
                    print(f"[DEBUG] Selected Arabic voice: {voice.name}")
                    break
                elif self.current_lang == "en" and any('en' in l or 'english' in l for l in lang_lower):
                    local_tts.setProperty('voice', voice.id)
                    selected = True
                    print(f"[DEBUG] Selected English voice: {voice.name}")
                    break

            if not selected and voices:
                local_tts.setProperty('voice', voices[0].id)  # Force first voice
                print(f"[WARN] No suitable voice for {self.current_lang.upper()} found, using default: {voices[0].name}")

            print(f"[DEBUG] Speaking: '{text}' with voice rate 160, volume 1.0")
            local_tts.say(text)
            local_tts.runAndWait()
            local_tts.stop()
            print("[DEBUG] TTS completed")
        except Exception as e:
            print(f"[speak_sync] error: {e}")

    def speak_async(self, text):
        """Run speak_sync in a background thread."""
        threading.Thread(target=lambda: self.speak_sync(text), daemon=True).start()

    def get_field_label(self, field_id):
        """Get label text for question based on field_id."""
        if field_id.startswith("header_"):
            key = field_id[7:]  # remove "header_"
            return self.translations[self.current_lang].get(key, key)
        elif field_id.startswith("table_row_"):
            parts = field_id.split("_")
            if len(parts) >= 4 and parts[:2] == ['table', 'row']:
                field = "_".join(parts[3:])
                return self.translations[self.current_lang]["table_headers"].get(field, field)
        return field_id

    def create_ui(self):
        """Build UI mimicking the Excel sheet layout."""
        # Main scrollable frame for overall content
        main_scroll = ctk.CTkScrollableFrame(self.root, corner_radius=10)
        main_scroll.pack(fill="both", expand=True, padx=10, pady=10)

        # Header main frame - transparent
        header_main = ctk.CTkFrame(main_scroll, fg_color="transparent")
        header_main.pack(fill="x", pady=10)

        t = self.translations[self.current_lang]

        # Logo left - fixed width
        logo_fr = ctk.CTkFrame(header_main, fg_color="transparent", width=150)
        logo_fr.pack(side="left", fill="y", padx=(50, 30))
        logo_fr.grid_propagate(False)  # Fix width
        try:
            logo_img = ctk.CTkImage(light_image=Image.open("logo.png"), size=(120, 100))
            logo_lbl = ctk.CTkLabel(logo_fr, image=logo_img, text="")
            logo_lbl.pack(pady=10)
        except Exception as e:
            print(f"Logo load error: {e}")
            logo_lbl = ctk.CTkLabel(logo_fr, text="LOGO", font=ctk.CTkFont(size=12, weight="bold"))
            logo_lbl.pack(pady=10)

        # Right side: Orange block for company name
        right_header = ctk.CTkFrame(header_main, fg_color="transparent")
        right_header.pack(side="right", fill="both", expand=True)
        
        # Add top spacer for vertical centering (adjust height as needed, e.g., 20 to shift down)
        top_spacer = ctk.CTkFrame(right_header, fg_color="transparent", height=20)
        top_spacer.pack(fill="x")
        top_spacer.grid_propagate(False)
        
        orange_block = ctk.CTkFrame(right_header, fg_color="#FF8C00", height=80, corner_radius=0)
        orange_block.pack(fill="x")
        orange_block.grid_propagate(False)
        company_lbl = ctk.CTkLabel(
            orange_block,
            text=t["title"],
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="white"
        )
        company_lbl.pack(expand=True, pady=10)

        # Below header_main: Blue block for form title (full width)
        form_block = ctk.CTkFrame(main_scroll, fg_color="#000080", height=40, corner_radius=0)
        form_block.pack(fill="x", pady=(0, 10))
        form_block.grid_propagate(False)
        form_lbl = ctk.CTkLabel(
            form_block,
            text=t["form_title"],
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="white"
        )
        form_lbl.pack(expand=True, pady=10)

        # Fields block - light blue divided header block
        fields_block = ctk.CTkFrame(main_scroll, fg_color="#E6F3FF", corner_radius=10, border_width=2, border_color="#000080")
        fields_block.pack(fill="x", pady=5)

        # Status label in fields_block bottom right
        self.status_label = ctk.CTkLabel(fields_block, text="", width=200, anchor="e", text_color="navy")
        self.status_label.pack(side="bottom", fill="x", padx=10, pady=5)    

        # Header grid inside fields_block
        header_grid = ctk.CTkFrame(fields_block, fg_color="transparent")
        header_grid.pack(padx=20, pady=10, fill="x")
        for i in range(6):
            header_grid.grid_columnconfigure(i, weight=1)

        # Row 1: Contract No., Contract Title, Report No., Activity Date
        ctk.CTkLabel(header_grid, text=t["contract_number"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.header_entries["contract_number"] = self.create_entry_with_click_voice(header_grid, "header_contract_number", row=0, col=1, pady=5)
        ctk.CTkLabel(header_grid, text=t["contract_title"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=0, column=2, sticky="w", padx=5, pady=5)
        self.header_entries["contract_title"] = self.create_entry_with_click_voice(header_grid, "header_contract_title", row=0, col=3, pady=5)
        ctk.CTkLabel(header_grid, text=t["report_number"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=0, column=4, sticky="w", padx=5, pady=5)
        self.header_entries["report_number"] = self.create_entry_with_click_voice(header_grid, "header_report_number", row=0, col=5, pady=5)

        # Row 2: Activity Date (shifted to row 1 col 5? but keep as is)
        # In image, Activity Date is in first row right
        # Adjust: put date in row=0 col=5
        # Wait, in code above, report_no col4-5, but add date to row=0 col=5? No, make col=6 if needed, but for now, extend columns if necessary.
        # To match, perhaps 4 sections: left contract/po/drawing/job, middle title/wps/project/line, right report/date/site/location
        # But current is fine, adjust pady for spacing.

        ctk.CTkLabel(header_grid, text=t["date"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        date_entry = ctk.CTkEntry(header_grid, width=150, fg_color="white")
        date_entry.insert(0, self.header_data["date"])
        date_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        self.header_entries["date"] = date_entry

        # Row 3: PO / WO No., Client WPS No., Project Title/Well ID
        ctk.CTkLabel(header_grid, text=t["po_wo_number"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.header_entries["po_wo_number"] = self.create_entry_with_click_voice(header_grid, "header_po_wo_number", row=2, col=1, pady=5)
        ctk.CTkLabel(header_grid, text=t["client_wps_number"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=2, column=2, sticky="w", padx=5, pady=5)
        self.header_entries["client_wps_number"] = self.create_entry_with_click_voice(header_grid, "header_client_wps_number", row=2, col=3, pady=5)
        ctk.CTkLabel(header_grid, text=t["project_title_wellID"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=2, column=4, sticky="w", padx=5, pady=5)
        self.header_entries["project_title_wellID"] = self.create_entry_with_click_voice(header_grid, "header_project_title_wellID", row=2, col=5, pady=5)

        # Row 4: Drawing/ISO No., Line No., Site Name
        ctk.CTkLabel(header_grid, text=t["drawing_no"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.header_entries["drawing_no"] = self.create_entry_with_click_voice(header_grid, "header_drawing_no", row=3, col=1, pady=5)
        ctk.CTkLabel(header_grid, text=t["line_no"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=3, column=2, sticky="w", padx=5, pady=5)
        self.header_entries["line_no"] = self.create_entry_with_click_voice(header_grid, "header_line_no", row=3, col=3, pady=5)
        ctk.CTkLabel(header_grid, text=t["site_name"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=3, column=4, sticky="w", padx=5, pady=5)
        self.header_entries["site_name"] = self.create_entry_with_click_voice(header_grid, "header_site_name", row=3, col=5, pady=5)

        # Row 5: Job Description (span 2 cols), Location (span 2 cols)
        ctk.CTkLabel(header_grid, text=t["job_desc"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=4, column=0, sticky="w", padx=5, pady=5)
        self.header_entries["job_desc"] = self.create_entry_with_click_voice(header_grid, "header_job_desc", row=4, col=1, colspan=2, pady=5)
        ctk.CTkLabel(header_grid, text=t["location"], font=ctk.CTkFont(weight="bold"), text_color="navy").grid(row=4, column=3, sticky="w", padx=5, pady=5)
        self.header_entries["location"] = self.create_entry_with_click_voice(header_grid, "header_location", row=4, col=4, colspan=2, pady=5)

        # Table section - Using a scrollable frame for vertical, canvas for horizontal
        table_outer = ctk.CTkFrame(main_scroll, corner_radius=10)
        table_outer.pack(fill="both", expand=True, pady=10)

        # Horizontal scroll canvas
        self.canvas = tk.Canvas(table_outer)
        h_scrollbar = ctk.CTkScrollbar(table_outer, orientation="horizontal", command=self.canvas.xview)
        h_scrollbar.pack(side="bottom", fill="x")
        self.canvas.configure(xscrollcommand=h_scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)

        # Vertical scrollbar
        v_scrollbar = ctk.CTkScrollbar(table_outer, orientation="vertical", command=self.canvas.yview)
        v_scrollbar.pack(side="right", fill="y")
        self.canvas.configure(yscrollcommand=v_scrollbar.set)

        # Inner frame
        inner_table = ctk.CTkFrame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=inner_table, anchor="nw")

        def update_scrollregion(event):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

        inner_table.bind("<Configure>", update_scrollregion)

        def update_canvas_width(event):
            canvas_width = event.width
            self.canvas.itemconfig(self.inner_id, width=canvas_width)

        self.canvas.bind("<Configure>", update_canvas_width)

        # Mouse wheel bindings for better scrolling
        def on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def on_shift_mousewheel(event):
            self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

        # Bind to root for global scrolling
        self.root.bind_all("<MouseWheel>", on_mousewheel)
        self.root.bind_all("<Shift-MouseWheel>", on_shift_mousewheel)
        # For Linux
        self.root.bind_all("<Button-4>", lambda e: self.canvas.yview_scroll(-1, "units"))
        self.root.bind_all("<Button-5>", lambda e: self.canvas.yview_scroll(1, "units"))
        self.root.bind_all("<Shift-Button-4>", lambda e: self.canvas.xview_scroll(-1, "units"))
        self.root.bind_all("<Shift-Button-5>", lambda e: self.canvas.xview_scroll(1, "units"))

        # Bind to canvas too
        self.canvas.bind("<MouseWheel>", on_mousewheel)
        self.canvas.bind("<Shift-MouseWheel>", on_shift_mousewheel)

        # Table headers - gray background like Excel
        headers_fr = ctk.CTkFrame(inner_table, fg_color=("gray85", "gray15"))
        headers_fr.pack(fill="x")

        header_fields = ["sr_no", "kp_sec", "weld_id", "wps_no", "material_gr_heat", "size", "thk", "weld_side", "welder_process", "visual_i", "visual_ii", "root_hot", "fill1", "fill2", "cap", "final", "fit_up", "mtrl_comb", "pipe_line", "pipe_no", "pipe_length", "remarks"]
        num_cols = len(header_fields)
        for i in range(num_cols):
            headers_fr.grid_columnconfigure(i, weight=1, minsize=70)  # Increased minsize for visibility
        
        for col, field in enumerate(header_fields):
            lbl = ctk.CTkLabel(headers_fr, text=t["table_headers"][field], font=ctk.CTkFont(weight="bold"), text_color=("black", "white"))
            lbl.grid(row=0, column=col, padx=1, pady=5, sticky="ew")

        # 10 data rows
        # 10 data rows
        self.table_entries = {}
        for row in range(1, 11):
            row_fr = ctk.CTkFrame(inner_table)
            row_fr.pack(fill="x")
            for i in range(num_cols):
                row_fr.grid_columnconfigure(i, weight=1, minsize=70)

            self.table_entries[row] = {}
            for col, field in enumerate(header_fields):
                if field == "sr_no":
                    sr_lbl = ctk.CTkLabel(row_fr, text=str(row), width=50, font=ctk.CTkFont(weight="bold"))
                    sr_lbl.grid(row=0, column=col, padx=1, pady=1, sticky="ew")
                else:
                    entry = self.create_entry_with_click_voice(row_fr, f"table_row_{row}_{field}", row=0, col=col)
                    self.table_entries[row][field] = entry

        # Welding Consumable section - light gray
        consumable_fr = ctk.CTkFrame(main_scroll, fg_color=("lightgray", "gray30"), corner_radius=5)
        consumable_fr.pack(fill="x", pady=10)
        for i in range(2):
            consumable_fr.grid_columnconfigure(i, weight=1)

        ctk.CTkLabel(consumable_fr, text=t["welding_consumable"], font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(consumable_fr, text=t["aws_classification"], font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.header_entries["aws_classification"] = self.create_entry_with_click_voice(consumable_fr, "header_aws_classification", row=1, col=1)
        ctk.CTkLabel(consumable_fr, text=t["electrode_dia"], font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, sticky="w", padx=5, pady=2)
        self.header_entries["electrode_dia"] = self.create_entry_with_click_voice(consumable_fr, "header_electrode_dia", row=2, col=1)
        ctk.CTkLabel(consumable_fr, text=t["manufacturer_batch"], font=ctk.CTkFont(weight="bold")).grid(row=3, column=0, sticky="w", padx=5, pady=2)
        self.header_entries["manufacturer_batch"] = self.create_entry_with_click_voice(consumable_fr, "header_manufacturer_batch", row=3, col=1)

        # Legends - light gray background
        legend_fr = ctk.CTkFrame(main_scroll, fg_color=("lightgray", "gray30"), corner_radius=5)
        legend_fr.pack(fill="x", pady=10)
        ctk.CTkLabel(legend_fr, text=t["material_grade_legend"], justify="left", anchor="w", font=ctk.CTkFont(size=10)).pack(pady=5, padx=5)
        ctk.CTkLabel(legend_fr, text=t["welding_process_legend"], justify="left", anchor="w", font=ctk.CTkFont(size=10)).pack(pady=5, padx=5)

        # Signatures frame - light blue like header
        sig_fr = ctk.CTkFrame(main_scroll, fg_color=("lightblue", "darkblue"), corner_radius=5)
        sig_fr.pack(fill="x", pady=10)
        for i in range(13):  # Enough columns
            sig_fr.grid_columnconfigure(i, weight=1)

        sig_keys = ["permit_holder", "qci", "pdo", "data_entry"]
        for i, key in enumerate(sig_keys):
            col_offset = i * 3
            ctk.CTkLabel(sig_fr, text=t[key], font=ctk.CTkFont(weight="bold", size=12)).grid(row=0, column=col_offset, sticky="w", padx=5, pady=5)
            ctk.CTkLabel(sig_fr, text=t["name"], font=ctk.CTkFont(weight="bold")).grid(row=1, column=col_offset, sticky="w", padx=5)
            self.signature_entries[f"{key}_name"] = ctk.CTkEntry(sig_fr, width=150, fg_color="white")
            self.signature_entries[f"{key}_name"].grid(row=1, column=col_offset+1, sticky="ew", padx=5, pady=2)
            ctk.CTkLabel(sig_fr, text=t["signature"], font=ctk.CTkFont(weight="bold")).grid(row=2, column=col_offset, sticky="w", padx=5)
            sig_entry = ctk.CTkEntry(sig_fr, placeholder_text="Signature", width=150, fg_color="white")
            self.signature_entries[f"{key}_signature"] = sig_entry
            sig_entry.grid(row=2, column=col_offset+1, sticky="ew", padx=5, pady=2)
            ctk.CTkLabel(sig_fr, text=t["date"], font=ctk.CTkFont(weight="bold")).grid(row=3, column=col_offset, sticky="w", padx=5)
            self.signature_entries[f"{key}_date"] = ctk.CTkEntry(sig_fr, width=150, fg_color="white")
            self.signature_entries[f"{key}_date"].insert(0, datetime.now().strftime("%Y-%m-%d"))
            self.signature_entries[f"{key}_date"].grid(row=3, column=col_offset+1, sticky="ew", padx=5, pady=2)

        # Buttons
        btn_fr = ctk.CTkFrame(main_scroll)
        btn_fr.pack(fill="x", pady=10)
        self.clear_btn = ctk.CTkButton(btn_fr, text=t["clear_form"], fg_color="gray", command=self.clear_form)
        self.clear_btn.pack(side="left", padx=5)
        self.export_btn = ctk.CTkButton(btn_fr, text=t["download_excel"], command=self.export_excel)
        self.export_btn.pack(side="right", padx=5)

    def create_entry_with_click_voice(self, parent, field_id, row=0, col=0, colspan=1, pady=2):
        """Create an entry that triggers voice input on click (no mic button)."""
        fr = ctk.CTkFrame(parent)
        fr.grid(row=row, column=col, columnspan=colspan, sticky="ew", padx=1, pady=pady)
        fr.grid_columnconfigure(0, weight=1)

        entry = ctk.CTkEntry(fr, fg_color="white")
        entry.grid(row=0, column=0, sticky="ew", padx=2, pady=2)
        # Bind left-click to trigger voice recording
        entry.bind("<Button-1>", lambda e: self.record_voice(field_id))

        # Store reference to entry only (no mic)
        self.entry_frames[field_id] = fr  # Keep for potential future use, but not mic

        print(f"[DEBUG] created click-to-voice entry for {field_id}")

        return entry

    # Update language (update labels)
    def update_language(self):
        t = self.translations[self.current_lang]
        # For simplicity, UI labels would need to be updated dynamically in a full implementation
        pass

    # Load saved data to UI
    def load_to_ui(self):
        # Load header
        for key, value in self.header_data.items():
            if key in self.header_entries:
                self.header_entries[key].delete(0, "end")
                self.header_entries[key].insert(0, value)

        # Load table
        for i, record in enumerate(self.records[:10]):
            row = i + 1
            for field, value in record.items():
                if field in self.table_entries[row]:
                    self.table_entries[row][field].delete(0, "end")
                    self.table_entries[row][field].insert(0, value)

        # Load signatures
        for k, v in self.signature_entries.items():
            v.delete(0, "end")
            v.insert(0, self.header_data.get(k, ""))

    # Save UI to data
    def save_from_ui(self):
        # Header
        for key, entry in self.header_entries.items():
            self.header_data[key] = entry.get().strip()

        # Table
        self.records = []
        for row in range(1, 11):
            row_data = {}
            for field, entry in self.table_entries[row].items():
                row_data[field] = entry.get().strip()
            if any(row_data.values()):  # If row not empty
                self.records.append(row_data)

        # Signatures
        for key, entry in self.signature_entries.items():
            self.header_data[key] = entry.get().strip()

        self.save_data()

    def clear_form(self):
        for entry in self.header_entries.values():
            entry.delete(0, "end")
        self.header_entries["date"].insert(0, datetime.now().strftime("%Y-%m-%d"))
        for row_entries in self.table_entries.values():
            for entry in row_entries.values():
                entry.delete(0, "end")
        for entry in self.signature_entries.values():
            entry.delete(0, "end")
        for key in self.signature_entries:
            if "_date" in key:
                entry = self.signature_entries[key]
                entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        # Re-enable mics if cleared
        for field_id in list(self.mic_buttons.keys()):
            self._clear_field(field_id)

    def record_voice(self, field_id):
        """Kick off TTS prompt then start background recording/transcription."""
        # Build human question using existing helper (get_field_label)
        try:
            field_label_text = self.get_field_label(field_id) or field_id
        except Exception:
            field_label_text = field_id

        if self.current_lang == "ar":
            question = f"ما هو {field_label_text.replace(':', '')}؟"
        else:
            question = f"What is your {field_label_text.replace(':', '').lower()}?"

        # Show question to user
        try:
            self.status_label.configure(text=question)
            self.root.update_idletasks()
        except Exception:
            pass

        # Speak synchronously so TTS finishes before recording
        try:
            self.speak_sync(question)
        except Exception as e:
            print("[record_voice] speak error:", e)

        # small pause
        time.sleep(0.15)

        # status update
        try:
            self.status_label.configure(text=self.translations[self.current_lang].get("recording", "Recording..."))
            self.root.update_idletasks()
        except Exception:
            pass

        # start background recording/transcription
        threading.Thread(target=self._record_audio, args=(field_id,), daemon=True).start()

    def _record_audio(self, field_id, timeout_seconds=4.5):
        """Record audio to a temp wav, transcribe (if model available), insert into the matching entry,
        and then start confirmation."""
        if self.is_recording:
            return
        self.is_recording = True
        audio = None
        stream = None
        tmp_path = None
        recognized_text = ""
        try:
            audio = pyaudio.PyAudio()
            with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as t:
                tmp_path = t.name

            stream = audio.open(format=pyaudio.paInt16, channels=1, rate=16000, input=True, frames_per_buffer=4000)
            frames = []
            start = time.time()
            while time.time() - start < timeout_seconds:
                try:
                    data = stream.read(4000, exception_on_overflow=False)
                    frames.append(data)
                except Exception as e:
                    # keep going if occasional read fails
                    print("[_record_audio] read error:", e)
                    continue

            wf = wave.open(tmp_path, "wb")
            wf.setnchannels(1)
            wf.setsampwidth(audio.get_sample_size(pyaudio.paInt16))
            wf.setframerate(16000)
            wf.writeframes(b"".join(frames))
            wf.close()

            # transcribe if whisper loaded
            if self.whisper_model:
                try:
                    lang_code = "ar" if self.current_lang == "ar" else "en"
                    result = self.whisper_model.transcribe(tmp_path, language=lang_code)
                    recognized_text = result.get("text", "").strip()
                except Exception as e:
                    print("[_record_audio] whisper error:", e)
                    recognized_text = ""
            else:
                # whisper unavailable -> inform user (no decision)
                recognized_text = ""
                print("[_record_audio] Warning: Whisper model not loaded; transcription unavailable.")

            # clean text: keep hyphens, keep digits' decimals, convert the word 'dash' to '-'
            if recognized_text:
                clean_text = re.sub(r"(?<!\d)\.(?!\d)", "", recognized_text)   # remove stray dots not part of decimals
                clean_text = re.sub(r"[^\w.\s-]", "", clean_text)              # allow letters, digits, dots, whitespace and hyphen
                clean_text = re.sub(r"\s+", " ", clean_text).strip()
                clean_text = re.sub(r'\bdash\b', '-', clean_text, flags=re.IGNORECASE)
            else:
                clean_text = ""

            # insert the tentative transcription into the UI (main thread)
            if clean_text:
                self.root.after(0, lambda fid=field_id, txt=clean_text: self._insert_text_to_field(fid, txt))
                # start confirmation in background
                threading.Thread(target=lambda: self._voice_confirm(field_id, clean_text), daemon=True).start()
            else:
                # No transcription available: fallback to GUI notice
                self.root.after(0, lambda: messagebox.showinfo("Info", "No speech recognized / speech-to-text unavailable"))

        except Exception as e:
            print("[_record_audio] error:", e)
        finally:
            try:
                if stream:
                    stream.stop_stream()
                    stream.close()
                if audio:
                    audio.terminate()
            except Exception:
                pass
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
            self.is_recording = False
            try:
                self.root.after(0, lambda: self.status_label.configure(text=""))
            except Exception:
                pass
    
    def get_entry_by_id(self, field_id):
        if field_id.startswith("header_"):
            return self.header_entries.get(field_id[7:], None)
        elif field_id.startswith("table_row_"):
            # expected format: "table_row_{row}_{field_name...}"
            parts = field_id.split("_")
            if len(parts) >= 4 and parts[:2] == ['table', 'row']:
                try:
                    row = int(parts[2])
                except Exception:
                    return None
                field = "_".join(parts[3:])
                return self.table_entries.get(row, {}).get(field, None)
        return None

    def _voice_confirm(self, field_id, recognized_text, max_retries=1):
        """Ask for voice confirmation (yes/no). On yes -> lock field (replace mic with label).
        On no -> clear and re-enable mic. falls back to GUI confirm if voice fails."""
        # speak confirmation
        try:
            if self.current_lang == "ar":
                confirm_prompt = self.translations["ar"].get("confirm_prompt", "قلت {}. هل تؤكد؟").format(recognized_text)
            else:
                confirm_prompt = self.translations["en"].get("confirm_prompt", "You said {}. Do you confirm?").format(recognized_text)
            self.speak_sync(confirm_prompt)
        except Exception as e:
            print("[_voice_confirm] TTS error:", e)

        time.sleep(0.2)

        yes_set_en = {"yes", "yeah", "yup", "yep", "confirm", "correct"}
        no_set_en = {"no", "nah", "nope", "incorrect", "wrong"}
        yes_set_ar = {"نعم", "ايوه", "ايه", "نَعَم"}
        no_set_ar = {"لا", "لأ", "لاا"}

        attempt = 0
        while attempt <= max_retries:
            attempt += 1
            # record short reply
            audio = None
            stream = None
            tmp = None
            response_text = ""
            try:
                audio = pyaudio.PyAudio()
                with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as t:
                    tmp = t.name
                stream = audio.open(format=pyaudio.paInt16, channels=1, rate=16000, input=True, frames_per_buffer=4000)
                # warm-up
                try:
                    for _ in range(2):
                        stream.read(4000, exception_on_overflow=False)
                except Exception:
                    pass
                frames = []
                start = time.time()
                rec_secs = 2.5
                while time.time() - start < rec_secs:
                    try:
                        data = stream.read(4000, exception_on_overflow=False)
                        frames.append(data)
                    except Exception:
                        pass
                wf = wave.open(tmp, "wb")
                wf.setnchannels(1)
                wf.setsampwidth(audio.get_sample_size(pyaudio.paInt16))
                wf.setframerate(16000)
                wf.writeframes(b"".join(frames))
                wf.close()
                # transcribe
                if self.whisper_model:
                    try:
                        lang_code = "ar" if self.current_lang == "ar" else "en"
                        res = self.whisper_model.transcribe(tmp, language=lang_code)
                        response_text = res.get("text", "").strip().lower()
                    except Exception as e:
                        print("[_voice_confirm] whisper confirm err:", e)
                        response_text = ""
                else:
                    response_text = ""
                if tmp and os.path.exists(tmp):
                    try:
                        os.unlink(tmp)
                    except Exception:
                        pass

                if not response_text:
                    if attempt <= max_retries:
                        self.speak_async(self.translations[self.current_lang].get("no_hear", "I didn't hear you. Please say yes or no."))
                        time.sleep(0.15)
                        continue
                    else:
                        # fallback to GUI confirm
                        break

                # tokenise (support Arabic unicode range too)
                tokens = set(re.split(r"\s+|[^\w\u0600-\u06FF]+", response_text))
                if (tokens & yes_set_en) or (tokens & yes_set_ar):
                    # confirmed -> lock field (main thread)
                    self.root.after(0, lambda fid=field_id: self._lock_field(fid))
                    self.root.after(0, lambda: self.status_label.configure(text="Confirmed and locked!"))
                    self.root.after(2000, lambda: self.status_label.configure(text=""))
                    return
                if (tokens & no_set_en) or (tokens & no_set_ar):
                    # rejected -> clear and recreate mic
                    self.speak_async(self.translations[self.current_lang].get("okay_retry", "Okay, please say it again."))
                    self.root.after(0, lambda fid=field_id: self._clear_field(fid))
                    self.root.after(0, lambda: self.status_label.configure(text="Cleared - please re-record"))
                    self.root.after(2000, lambda: self.status_label.configure(text=""))
                    return

                # else unrecognized -> retry if attempts left
                if attempt <= max_retries:
                    self.speak_async(self.translations[self.current_lang].get("no_catch", "I didn't catch that. Please say yes or no."))
                    time.sleep(0.15)
                    continue
                else:
                    break

            except Exception as e:
                print("[_voice_confirm] error:", e)
                try:
                    if stream:
                        stream.stop_stream()
                        stream.close()
                    if audio:
                        audio.terminate()
                except Exception:
                    pass
                time.sleep(0.1)
                continue
            finally:
                try:
                    if stream:
                        stream.stop_stream()
                        stream.close()
                    if audio:
                        audio.terminate()
                except Exception:
                    pass

        # fallback GUI confirm (runs on main thread)
        try:
            t = self.translations[self.current_lang]
            confirm_msg = t.get("confirm_text", "Is this correct: '{}'?").format(recognized_text)
            if messagebox.askyesno("Confirm Input", confirm_msg):
                self._lock_field(field_id)
                self.status_label.configure(text="Confirmed and locked!")
                self.root.after(2000, lambda: self.status_label.configure(text=""))
            else:
                self._clear_field(field_id)
                self.status_label.configure(text="Cleared - please re-record")
                self.root.after(2000, lambda: self.status_label.configure(text=""))
        except Exception as e:
            print("[_voice_confirm] fallback GUI error:", e)

    def _insert_text_to_field(self, field_id, value):
        entry = self.get_entry_by_id(field_id)
        if entry:
            entry.delete(0, "end")
            entry.insert(0, value)

    def _lock_field(self, field_id):
        """Lock the field (make readonly/disabled) and remove its mic button."""
        entry_widget = self.get_entry_by_id(field_id)
        if entry_widget:
            try:
                entry_widget.configure(state="readonly")
            except Exception:
                try:
                    entry_widget.configure(state="disabled")
                except Exception:
                    pass

        # destroy mic if exists
        mic = self.mic_buttons.get(field_id)
        if mic:
            try:
                mic.destroy()
            except Exception:
                pass
            # remove from registry
            try:
                del self.mic_buttons[field_id]
            except KeyError:
                pass

        print(f"[DEBUG] Locked and removed mic for '{field_id}'")

    def _clear_field(self, field_id):
        """Clear entry and recreate mic button."""
        entry = self.get_entry_by_id(field_id)
        if not entry:
            return

        # Clear entry
        try:
            entry.configure(state="normal")
            entry.delete(0, "end")
        except Exception:
            try:
                entry.configure(state="normal")
                entry.delete("1.0", "end")
            except Exception:
                pass

        # Recreate mic in container
        container = self.entry_frames.get(field_id)
        if container:
            # destroy old mic if any
            old_mic = self.mic_buttons.get(field_id)
            if old_mic:
                try:
                    old_mic.destroy()
                except Exception:
                    pass
                try:
                    del self.mic_buttons[field_id]
                except KeyError:
                    pass
            # create new mic button and store it
            mic = ctk.CTkButton(container, text="🎤", width=30, height=25, corner_radius=8, command=lambda f=field_id: self.record_voice(f))
            mic.grid(row=0, column=1, sticky="e", padx=(2, 2), pady=2)
            self.mic_buttons[field_id] = mic

    def export_excel(self):
        self.save_from_ui()  # Save current UI to data
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Daily Welding Report"

            # Add styles
            header_font = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            t = self.translations[self.current_lang]

            # Title
            ws['A1'] = t["title"]
            ws['A1'].font = Font(size=18, bold=True)
            ws.merge_cells('A1:F1')

            # Form title
            ws['A2'] = t["form_title"]
            ws['A2'].font = Font(size=16, bold=True)
            ws.merge_cells('A2:F2')

            # Header fields (approximate Excel positions)
            row = 4
            ws[f'A{row}'] = t["contract_number"]
            ws[f'B{row}'] = self.header_data["contract_number"]
            ws[f'G{row}'] = t["contract_title"]
            ws[f'H{row}'] = self.header_data["contract_title"]
            ws[f'M{row}'] = t["report_number"]
            ws[f'N{row}'] = self.header_data["report_number"]
            ws[f'S{row}'] = t["date"]
            ws[f'T{row}'] = self.header_data["date"]
            row += 1

            ws[f'A{row}'] = t["po_wo_number"]
            ws[f'B{row}'] = self.header_data["po_wo_number"]
            ws[f'G{row}'] = t["client_wps_number"]
            ws[f'H{row}'] = self.header_data["client_wps_number"]
            ws[f'M{row}'] = t["project_title_wellID"]
            ws[f'N{row}'] = self.header_data["project_title_wellID"]
            row += 1

            ws[f'A{row}'] = t["drawing_no"]
            ws[f'B{row}'] = self.header_data["drawing_no"]
            ws[f'G{row}'] = t["line_no"]
            ws[f'H{row}'] = self.header_data["line_no"]
            ws[f'M{row}'] = t["site_name"]
            ws[f'N{row}'] = self.header_data["site_name"]
            row += 1

            ws[f'A{row}'] = t["job_desc"]
            ws[f'B{row}'] = self.header_data["job_desc"]
            ws.merge_cells(f'B{row}:C{row}')
            ws[f'D{row}'] = t["location"]
            ws[f'E{row}'] = self.header_data["location"]
            ws.merge_cells(f'E{row}:F{row}')
            row += 3

            # Table headers
            headers = list(t["table_headers"].values())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.border = thin_border

            # Table data
            for r, record in enumerate(self.records, row + 1):
                for col, (field, value) in enumerate(record.items(), 1):
                    cell = ws.cell(row=r, column=col, value=value)
                    cell.border = thin_border

            # Consumables
            row += len(self.records) + 2
            ws[f'A{row}'] = t["welding_consumable"]
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = t["aws_classification"]
            ws[f'B{row}'] = self.header_data["aws_classification"]
            row += 1
            ws[f'A{row}'] = t["electrode_dia"]
            ws[f'B{row}'] = self.header_data["electrode_dia"]
            row += 1
            ws[f'A{row}'] = t["manufacturer_batch"]
            ws[f'B{row}'] = self.header_data["manufacturer_batch"]

            # Legends
            row += 2
            ws[f'A{row}'] = t["material_grade_legend"]
            ws[f'A{row}'].font = Font(size=8)
            row += 1
            ws[f'A{row}'] = t["welding_process_legend"]
            ws[f'A{row}'].font = Font(size=8)

            # Signatures
            row += 2
            sig_row = row
            sig_keys = ["permit_holder", "qci", "pdo", "data_entry"]
            for i, key in enumerate(sig_keys):
                col = i * 4 + 1
                ws[f'{chr(64 + col)}{sig_row}'] = t[key]
                ws[f'{chr(64 + col)}{sig_row}'].font = Font(bold=True)
                ws[f'{chr(64 + col)}{sig_row + 1}'] = t["name"]
                ws[f'{chr(64 + col + 1)}{sig_row + 1}'] = self.header_data[f"{key}_name"]
                ws[f'{chr(64 + col)}{sig_row + 2}'] = t["signature"]
                ws[f'{chr(64 + col + 1)}{sig_row + 2}'] = self.header_data[f"{key}_signature"]
                ws[f'{chr(64 + col)}{sig_row + 3}'] = t["date"]
                ws[f'{chr(64 + col + 1)}{sig_row + 3}'] = self.header_data[f"{key}_date"]

            filename = f"welding_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            messagebox.showinfo("Success", f"Report exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")

    def save_data(self):
        data = {
            "header": self.header_data,
            "records": self.records,
        }
        try:
            with open(self.data_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving data: {e}")

    def load_data(self):
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.header_data = data.get("header", self.header_data)
                    self.records = data.get("records", [])
            except:
                pass


if __name__ == "__main__":
    root = ctk.CTk()
    app = WeldingShopApp(root)
    root.mainloop()